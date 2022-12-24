from flask import Flask,render_template,request

app = Flask(__name__)

@app.route("/",methods=['GET','POST'])
def index():    
    return render_template('index.html')

@app.route('/name/<username>')
def show_user(username):
    return f"<h3>您的姓名是<strong>{username}</strong></h3>"

from openpyxl import load_workbook
@app.route('/login',methods=['GET','POST'])
def login():
    if request.method == "POST":
        print('由表單傳送過來的')
        print(f"email:{request.form['email']}")
        print(f"email:{request.form['password']}")
        wb = load_workbook('static/others/login.xlsx')
        login_sheet = wb['工作表1']
        for row in login_sheet.iter_rows():
            print(row[0].value)
            print(row[1].value)
    return render_template('login.html')